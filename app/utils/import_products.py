from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from slugify import slugify
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import AsyncSessionLocal
from app.models.brand import Brand
from app.models.category import Category
from app.models.product import Product


@dataclass
class ImportStats:
    brands_created: int = 0
    categories_created: int = 0
    products_created: int = 0
    products_updated: int = 0


def normalize_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_decimal(value: object | None) -> Decimal | None:
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(",", ".")
    return Decimal(text).quantize(Decimal("0.01"))


def split_multi_value_urls(value: object | None) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    normalized = text.replace("\n", ";")
    return [item.strip() for item in normalized.split(";") if item.strip()]


def status_to_flags(status: str | None) -> tuple[str | None, bool, int]:
    mapping = {
        "В наявності": ("in_stock", True, 1),
        "Немає в наявності": ("out_of_stock", False, 0),
    }
    return mapping.get(status or "", ("unknown", False, 0))


def build_product_slug(name: str, sku: str | None) -> str:
    base = slugify(name, lowercase=True)
    if sku:
        return f"{base}-{slugify(sku, lowercase=True)}"
    return base


def build_category_slug(path_parts: list[str]) -> str:
    return slugify("-".join(path_parts), lowercase=True)


def parse_xlsx(file_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed_rows: list[dict[str, object]] = []
    for row in rows[1:]:
        item = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if not normalize_text(item.get("Артикул")):
            continue
        parsed_rows.append(item)
    return parsed_rows


async def get_or_create_brand(session: AsyncSession, name: str, stats: ImportStats) -> Brand:
    result = await session.execute(select(Brand).where(Brand.name == name))
    brand = result.scalar_one_or_none()
    if brand:
        return brand

    brand = Brand(name=name, slug=slugify(name, lowercase=True))
    session.add(brand)
    await session.flush()
    stats.brands_created += 1
    return brand


async def get_or_create_category_tree(
    session: AsyncSession,
    category_path: str,
    cache: dict[str, Category],
    stats: ImportStats,
) -> Category | None:
    parts = [part.strip() for part in category_path.split("/") if part and part.strip()]
    if not parts:
        return None

    parent: Category | None = None
    seen_parts: list[str] = []
    for part in parts:
        seen_parts.append(part)
        key = "/".join(seen_parts)
        if key in cache:
            parent = cache[key]
            continue

        slug = build_category_slug(seen_parts)
        result = await session.execute(select(Category).where(Category.slug == slug))
        category = result.scalar_one_or_none()
        if not category:
            category = Category(
                name=part,
                slug=slug,
                is_active=True,
                parent_id=parent.id if parent else None,
            )
            session.add(category)
            await session.flush()
            stats.categories_created += 1
        cache[key] = category
        parent = category
    return parent


def build_attributes(row: dict[str, object]) -> dict:
    gallery = normalize_text(row.get("Галерея"))
    photos = split_multi_value_urls(row.get("Фото"))
    extra_categories = normalize_text(row.get("Дополнительные разделы"))
    attrs = {
        "parent_sku": normalize_text(row.get("Родительский артикул")),
        "display_sku": normalize_text(row.get("Артикул для отображения на сайте")),
        "size": normalize_text(row.get("Размер (UA)")),
        "color": normalize_text(row.get("Цвет")),
        "mpn": normalize_text(row.get("Код производителя товара (MPN)")),
        "source_url": normalize_text(row.get("Ссылка")),
        "image_urls": photos,
        "gallery": [item.strip() for item in gallery.split(";") if item.strip()] if gallery else [],
        "extra_category_paths": [item.strip() for item in extra_categories.split(";") if item.strip()] if extra_categories else [],
        "source_added_at": normalize_text(row.get("Дата добавления")),
    }
    return {key: value for key, value in attrs.items() if value not in (None, "", [], {})}


async def import_products(file_path: Path) -> ImportStats:
    rows = parse_xlsx(file_path)
    stats = ImportStats()
    category_cache: dict[str, Category] = {}

    async with AsyncSessionLocal() as session:
        assert isinstance(session, AsyncSession)

        for row in rows:
            sku = normalize_text(row.get("Артикул"))
            if not sku:
                continue

            brand_name = normalize_text(row.get("Бренд"))
            category_path = normalize_text(row.get("Раздел"))
            brand = await get_or_create_brand(session, brand_name, stats) if brand_name else None
            category = (
                await get_or_create_category_tree(session, category_path, category_cache, stats)
                if category_path
                else None
            )

            name = normalize_text(row.get("Название модификации (UA)")) or normalize_text(row.get("Название (UA)"))
            if not name:
                continue

            availability_status, is_active, stock_quantity = status_to_flags(normalize_text(row.get("Наличие")))
            description = normalize_text(row.get("Описание товара (UA)"))
            short_description = normalize_text(row.get("Короткое описание (UA)"))
            photo_urls = split_multi_value_urls(row.get("Фото"))
            image_url = photo_urls[0] if photo_urls else None
            external_url = normalize_text(row.get("Ссылка"))
            price = normalize_decimal(row.get("Цена"))
            recommended_retail_price = normalize_decimal(row.get("РРЦ"))
            attributes_json = build_attributes(row)

            result = await session.execute(select(Product).where(Product.sku == sku))
            product = result.scalar_one_or_none()
            payload = {
                "name": name,
                "slug": build_product_slug(name, sku),
                "description": description,
                "short_description": short_description,
                "price": price or Decimal("0.00"),
                "recommended_retail_price": recommended_retail_price,
                "sku": sku,
                "stock_quantity": stock_quantity,
                "is_active": is_active,
                "image_url": image_url,
                "external_url": external_url,
                "availability_status": availability_status,
                "attributes_json": attributes_json or None,
                "brand_id": brand.id if brand else None,
                "category_id": category.id if category else None,
            }

            if product:
                for key, value in payload.items():
                    setattr(product, key, value)
                stats.products_updated += 1
            else:
                session.add(Product(**payload))
                stats.products_created += 1

        await session.commit()

    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import products from xlsx")
    parser.add_argument("--file", required=True, help="Path to xlsx file visible inside the runtime")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    result = asyncio.run(import_products(Path(args.file)))
    print(
        "Import completed:",
        f"brands_created={result.brands_created}",
        f"categories_created={result.categories_created}",
        f"products_created={result.products_created}",
        f"products_updated={result.products_updated}",
    )
